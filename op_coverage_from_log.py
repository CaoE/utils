#-*- coding:utf-8 -*-

import pandas as pd 
import re 
from openpyxl import Workbook

def test_re(input_file):
    book = Workbook()
    sheet = book.active
    sheet.title = 'WW35' #给Excel标签命名为Fisrt

    Raw = 0

    f = open(input_file)
    lines = f.readlines() #读txt文档所有的数据
    i = 0
    preop = "fakeop"
    while i < len(lines):
        var = lines[i]
        if var.split(':')[0] == "# Benchmarking PyTorch":
            if var.split(':')[1][1:] != preop:
                preop = var.split(':')[1][1:]
                Raw += 1
            sheet.cell(Raw+1, 1).value = var.split(':')[1] # op name


        elif var.split(' ')[0] == "Forward" or var.split(' ')[0] == "Backward":
            lvar = lines[i-2] #name
            if lines[i+3].find('~~~~') != -1 or var.find('-1') != -1:
                errm = lines[i+2]
                if lvar.find('bfloat16') != -1:
                    if var.split(' ')[0] == "Forward":
                        if errm.find('BFloat16') != -1 or errm.find('float') != -1 or errm.find('double') != -1 or errm.find('Float') != -1 or errm.find('Double') != -1: 
                            sheet.cell(Raw+1, 4).value = 'N'
                        else:
                            sheet.cell(Raw+1, 4).value = 'need check'
                    elif var.split(' ')[0] == "Backward":
                        if errm.find('derivative') != -1 or errm.find('differentiable') != -1:
                            sheet.cell(Raw+1, 5).value = 'N/A'
                        elif errm.find('BFloat16') != -1 or errm.find('float') != -1 or errm.find('double') != -1 or errm.find('Float') != -1 or errm.find('Double') != -1:
                            sheet.cell(Raw+1, 5).value = 'N'
                        elif errm.find('check_uniform_bounds') != -1 or errm.find('does not require grad') != -1 or errm.find('object') != -1:
                            sheet.cell(Raw+1, 5).value = 'N/A'
                        else:
                            sheet.cell(Raw+1, 5).value = 'need check'
                else:
                    if var.split(' ')[0] == "Forward":
                        sheet.cell(Raw+1, 2).value = 'need check'
                    elif var.split(' ')[0] == "Backward":
                        if errm.find('derivative') != -1 or errm.find('differentiable') != -1:
                            sheet.cell(Raw+1, 3).value = 'N/A'
                        elif errm.find('check_uniform_bounds') != -1 or errm.find('does not require grad') != -1 or errm.find('object') != -1:
                            sheet.cell(Raw+1, 3).value = 'N/A'
                        else:
                            sheet.cell(Raw+1, 3).value = 'need check'
            elif lvar.split(':')[0] == "# Name":
                # if lvar.find('dtype_one') == -1 and lvar.find('dtype_two') == -1: # just dtype
                if lvar.find('bfloat16') != -1:
                    if var.split(' ')[0] == "Forward":
                        sheet.cell(Raw+1, 4).value = 'Y'
                    elif var.split(' ')[0] == "Backward":
                        sheet.cell(Raw+1, 5).value = 'Y'
                else:
                    if var.split(' ')[0] == "Forward":
                        sheet.cell(Raw+1, 2).value = 'Y'
                    elif var.split(' ')[0] == "Backward":
                        sheet.cell(Raw+1, 3).value = 'Y'
        
        
        i += 1
    book.save('internal_pytorch_all.xlsx')#保存为sample文档
    f.close()


if __name__ == "__main__":
    test_re("internal_pytorch_all.txt")

