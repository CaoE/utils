#-*- coding:utf-8 -*-
 
#安装好环境 pandas库，和openpyxl
import pandas as pd 
import re 
from openpyxl import Workbook

def texto_excel(input_file):
    book = Workbook()
    sheet = book.active
    sheet.title = 'First' #给Excel标签命名为Fisrt
    #book.create_sheet(title = 'sheet2',index = 1)  #新建一个sheet名为sheet2,在二号位
    Raw = 0
    Col = 0

    inputSameFlag = False
    f = open(input_file)
    lines = f.readlines() #读txt文档所有的数据
    i = 0
    preop = "fakeop"
    preinput = "fakeinput"
    preinput1 = "fakeinput1"
    preinput2 = "fakeinput2"
    forwardRaw = Raw
    backwardRaw = Raw
    prestatu = "Backward"
    while i < len(lines):
        var = lines[i]
        if var.split(':')[0] == "# Benchmarking PyTorch":
            if var.split(':')[1][1:] != preop:
                preop = var.split(':')[1][1:]
                Raw += 1
                if preop == "tensordot\n":
                    print(preop)
            sheet.cell(Raw+1, 1).value = var.split(':')[1] # op name
        
        elif var.split(':')[0] == "# Input":
            index = var.find('dtype')
            index1 = var.find('dtype_one')
            index2 = var.find('dtype_two')
            if index != -1:
                # if var.split(', dtype')[0] != preinput:
                # preinput = var.split(', dtype')[0]
                if lines[i+1].split(' ')[0] == "Forward":
                    if prestatu == "Backward":
                        prestatu = "Forward"
                        preinput = "fakeinput"
                        backwardRaw = Raw+1
                    if var.split(', dtype')[0] != preinput:    
                        Raw += 1
                        preinput = var.split(', dtype')[0]
                        sheet.cell(Raw+1, 2).value = var.split(', dtype')[0] # input shape
                elif lines[i+1].split(' ')[0] == "Backward":
                    if prestatu == "Forward":
                        Raw = backwardRaw
                        prestatu = "Backward"
                        preinput = var.split(', dtype')[0]
                    elif var.split(', dtype')[0] != preinput:
                        preinput = var.split(', dtype')[0]
                        Raw += 1

            elif index1 != -1 and index2 != -1:
                if lines[i+1].split(' ')[0] == "Forward":
                    Raw += 1
                    if prestatu == "Backward":
                        prestatu = "Forward"
                        backwardRaw = Raw
                    sheet.cell(Raw+1, 2).value = var.replace("device: cpu"," ") # input shape
                elif lines[i+1].split(' ')[0] == "Backward":
                    if prestatu == "Forward":
                        Raw = backwardRaw
                        prestatu = "Backward"
                    else:
                        Raw += 1


                # Raw += 1
                # sheet.cell(Raw+1, 2).value = var.replace("device: cpu"," ") # input shape
                # sheet.cell(Raw+1, 9).value = var.split(':')[1][index1:index2] # dtype1
                # sheet.cell(Raw+1, 10).value = var.split(':')[1][index2:] #dtype2
            

        elif var.split(' ')[0] == "Forward" or var.split(' ')[0] == "Backward":
            lvar = lines[i-2] #name
            if lvar.split(':')[0] == "# Name":
                if lvar.find('dtype_one') == -1 and lvar.find('dtype_two') == -1: # just dtype
                    if lvar.find('float32') != -1:
                        if var.split(' ')[0] == "Backward":
                            sheet.cell(Raw+1, 4).value = var.split(':')[1]
                        # elif lvar.split(':')[1][-4:] == "bwd1":
                        #     sheet.cell(Raw+1, 6).value = var.split(':')[1]
                        elif var.split(' ')[0] == "Forward":
                            sheet.cell(Raw+1, 3).value = var.split(':')[1]

                    elif lvar.find('bfloat16') != -1:
                        if var.split(' ')[0] == "Backward":
                            sheet.cell(Raw+1, 6).value = var.split(':')[1]
                        # elif lvar.split(':')[1][-4:] == "bwd1":
                        #     sheet.cell(Raw+1, 5).value = var.split(':')[1]
                        elif var.split(' ')[0] == "Forward":
                            sheet.cell(Raw+1, 5).value = var.split(':')[1]
                else:# dtype1 and dtype2
                    if var.split(' ')[0] == "Backward":
                        sheet.cell(Raw+1, 5).value = var.split(':')[1]
                    # elif lvar.split(':')[1][-4:] == "bwd1":
                    #     sheet.cell(Raw+1, 4).value = var.split(':')[1]
                    elif var.split(' ')[0] == "Forward":
                        sheet.cell(Raw+1, 3).value = var.split(':')[1]

            # sheet.cell(Raw+1, 4).value = var.split(' ')[0]
            # sheet.cell(Raw+1, 5).value = var.split(':')[1]
        i += 1
    book.save('inference_cpu_bs_ins0_internal_dev_in7528_iters1_cacheflush.xlsx')#保存为sample文档
    f.close()

if __name__ == "__main__":
    # replace_symbol(txt_path, txt1_path, Del_array, Aim_symbol )#处理原文本，生成一个匹配excel输入的文本
    # to_excel(txt1_path, Split_flag ) #将生成的文本以，分割；填入excel的每个表格
    texto_excel("inference_cpu_bs_ins0_internal_dev_in7528_iters1_cacheflush.txt")

