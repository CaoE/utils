#！/usr/bin/env python3
#-*- coding:utf-8 -*-
 
#安装好环境 pandas库，和openpyxl
import pandas as pd 
import re 
from openpyxl import Workbook
 
txt_path ='H:\\Python\\txt_process\\in.txt'  #源文档的文件路径
txt1_path ='H:\\Python\\txt_process\\out.txt' #数据基本处理之后生成txt文档路径
 
Del_array = ['{', '}', '//']
Aim_symbol = ''
 
Split_flag =','#以此符号分割每一行的文本
#替换Del_array中的内容为Aim_symbol
#para@--in_file_path----输入txt文档位置
#       out_file_path---替换之后输出的txt文档位置
#       del_array---源文本
#       _aim_symbol-替换成为的文本
def replace_symbol(in_file_path,out_file_path,del_array,_aim_symbol):
    infile = open(in_file_path)
    outfile = open(out_file_path, 'w') #在out_file_path新建一个文件
    outfile = open(out_file_path, 'r+') #打开此文件可写权限
    lenth = len(del_array)
    for line in infile.readlines():
        demo = line
        x = 0
        while x < len(del_array):
            demo = demo.replace(del_array[x], _aim_symbol) #循环替换del_array数组中的字符
            x = x + 1
        outfile.write(demo)#将替换后的字符串写入out_file
    infile.close()
    outfile.close() #类似于保存
#导出txt到Excel的每一列，以，为分割符
def to_excel(input_file, device_symbol):
    book = Workbook()
    sheet = book.active
    sheet.title = 'First' #给Excel标签命名为Fisrt
    #book.create_sheet(title = 'sheet2',index = 1)  #新建一个sheet名为sheet2,在二号位
    Raw = 0
    Col = 0
    f = open(input_file)
    lines = f.readlines() #读txt文档所有的数据
    for line in lines:
        value_count = len(line.split(device_symbol))
        for Col in range(0, value_count):
            sheet.cell(Raw+1, Col+1).value = line.split(device_symbol)[Col]#写入Excel每格数据，此数据以split_flag分割。
            Col = Col + 1
        Col = 0
        Raw = Raw+1
    book.save('sample.xlsx')#保存为sample文档
    f.close()

def texto_excel(input_file):
    book = Workbook()
    sheet = book.active
    sheet.title = 'First' #给Excel标签命名为Fisrt
    #book.create_sheet(title = 'sheet2',index = 1)  #新建一个sheet名为sheet2,在二号位
    Raw = 0
    Col = 0
    preflag = False
    f = open(input_file)
    lines = f.readlines() #读txt文档所有的数据
    i = 0
    while i < len(lines):
        var = lines[i]
        if var[0:3] == "pre":
            i = i + 5
            var = lines[i]
        elif var.split(':')[0] == "# Benchmarking PyTorch":
            sheet.cell(Raw+1, 1).value = var.split(':')[1]
            Raw += 1 
        elif var.split(':')[0] == "# Name":
            sheet.cell(Raw+1, 2).value = var.split(':')[1]
        elif var.split(':')[0] == "# Input":
            sheet.cell(Raw+1, 3).value = var[9:]
        elif var.split(' ')[0] == "Forward" or var.split(' ')[0] == "Backward":
            sheet.cell(Raw+1, 4).value = var.split(' ')[0]
            sheet.cell(Raw+1, 5).value = var.split(':')[1]
        i += 1

if __name__ == "__main__":
    # replace_symbol(txt_path, txt1_path, Del_array, Aim_symbol )#处理原文本，生成一个匹配excel输入的文本
    # to_excel(txt1_path, Split_flag ) #将生成的文本以，分割；填入excel的每个表格
    texto_excel("C:\\Files\\ssh_log\\[ssh ecao@10.239.60.4] (2021-04-26_090707).log")
